# -*- coding: utf-8 -*-
"""
Created on Tue Mar 23 14:53:44 2021

@author: Satgu
"""

from PyQt5.QtCore import pyqtSlot
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt,QTimer,QTime,QAbstractTableModel
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QTableView, QFileDialog,QMessageBox,QMainWindow, QLabel, QGridLayout,QDesktopWidget, QWidget,QTableWidget,QTableView,QTableWidgetItem,QHeaderView,QGraphicsScene,QGraphicsPixmapItem,QFileDialog
from covid_detection import Ui_Form
from PyQt5.QtWidgets import QMainWindow, QApplication, QLabel
from skimage.feature import (match_descriptors, corner_harris,corner_peaks, ORB, plot_matches)
from PyQt5.QtWidgets import QLabel, QVBoxLayout
import sys
import os
import cv2
import pandas as pd
import numpy as np
import matplotlib.image as mpimg
from PIL import Image
import xlsxwriter
import imutils
import matplotlib.pyplot as plt
from sklearn import datasets, svm, metrics
from sklearn.model_selection import train_test_split
from sklearn.model_selection import KFold 
from sklearn.metrics import plot_confusion_matrix
from sklearn.preprocessing import MinMaxScaler 
import matplotlib.pyplot as plt
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler 
from sklearn.metrics import accuracy_score
from sklearn.linear_model import LogisticRegression 
from sklearn.preprocessing import StandardScaler 
from sklearn.metrics import confusion_matrix, classification_report 
from sklearn.model_selection import train_test_split
from sklearn.model_selection import KFold 
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score
from sklearn.neighbors import KNeighborsClassifier
from sklearn.datasets import make_classification
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import f1_score
from sklearn.metrics import auc
from matplotlib import pyplot
from sklearn.datasets import make_classification
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import f1_score
from sklearn.metrics import auc
from sklearn.metrics import r2_score
from sklearn.metrics import mean_squared_error
from sklearn import metrics 
from matplotlib import pyplot
from sklearn.metrics import roc_curve
from sklearn.metrics import roc_auc_score
from PIL import Image
from keras.layers import Dense, Flatten
from keras import optimizers
from keras.models import Sequential
from keras.layers import Input, Lambda
from keras.models import Model
from keras.applications.vgg16 import VGG16
from keras.applications.vgg16 import preprocess_input
from keras.preprocessing.image import img_to_array
from keras.models import model_from_json
import tensorflow as tf
from tensorflow.keras.applications.resnet50 import ResNet50
from tensorflow.keras.applications import MobileNetV2
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Flatten
from tensorflow.keras.layers import Dense

from itertools import cycle
from tensorflow.keras.optimizers import RMSprop
from sklearn import svm, datasets
from sklearn.metrics import roc_curve, auc
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import label_binarize
from sklearn.multiclass import OneVsRestClassifier
from scipy import interp
from sklearn.metrics import roc_auc_score
import tensorflow as tf
import tensorflow_hub as hub
import keras
import glob,os, random
from keras.preprocessing.image import ImageDataGenerator
from keras.preprocessing import image
class MainWindow(QWidget,Ui_Form):
    
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)
   
        self.covid=0
        self.recovid=0
        self.dacovid=0
        self.renormal=0
        self.normal=0
        self.danormal=0
        self.covidlist=[]
        self.covidnumber=[]
        self.normallist=[]
        self.normalnumber=[]
        self.radioButton_6.setVisible(False)
        self.radioButton_7.setVisible(False)
        self.groupBox_5.setVisible(False)
        self.groupBox_16.setVisible(False)
        self.uploadataset.clicked.connect(self.uploaddataset)
        self.radioButton_6.toggled.connect(self.renamecovid)
        self.radioButton_7.toggled.connect(self.renamenormal)
        self.renameimages.clicked.connect(self.rename)
        self.listWidget.itemDoubleClicked.connect(self.doubleclickcovid)
        self.listWidget_2.itemDoubleClicked.connect(self.doubleclicknormal)
        self.convertexcel.clicked.connect(self.excelfile)
        self.ninfo.clicked.connect(self.ninformation)
        self.dataagumen.clicked.connect(self.datacogalt)
        self.datasplit.clicked.connect(self.datasplitfunction)
        self.kfoldshow.clicked.connect(self.kfoldyapalim)
        self.tableView_4.clicked.connect(self.test)
        self.tableView_6.clicked.connect(self.test)
        self.tlselect.clicked.connect(self.transferlearn)
        self.modeltest.clicked.connect(self.modeltestyap)
    def uploaddataset(self):
        self.foldername=[]
        folder1array=[]
        folder2array=[]
        self.path = QFileDialog().getExistingDirectory(None, 'Klasör seçiniz.')
        self.path=self.path+ '/'
        print(self.path)
        for i in os.listdir(self.path):
            self.foldername.append(i)
        self.folder1=self.path+self.foldername[0]+'/'
        self.folder2 = self.path+self.foldername[1]+'/'
        print(self.folder1)
        print(self.folder2)
        for i in os.listdir(self.folder1):
            folder1array.append(i)
        for i in os.listdir(self.folder2):
            folder2array.append(i)
       
        self.label.setText("Seçilen klasör:"+str(self.path[88:]))
        self.label_2.setText("Datasetin içinde bulunan klasör: "+str(self.foldername[0])+"  Görüntü sayısı:"+str(len(folder1array)))
        self.label_3.setText("Datasetin içinde bulunan klasör: "+str(self.foldername[1])+"  Görüntü sayısı:"+str(len(folder2array)))
        self.label_4.setText(str(self.foldername[0]))
        self.label_5.setText(str(self.foldername[1]))
        self.covidimages()
        self.normalimages()
        
    def covidimages(self):
      
        self.listWidget.clear()
        path= self.folder1
        for i in os.listdir(path):
              self.listWidget.insertItem(self.covid, i)
           
              self.covid = self.covid+1
        self.radioButton_6.setVisible(True)
        self.radioButton_6.setText("İsim değiştir")      	

              
    def normalimages(self):

        self.listWidget_2.clear()
        path= self.folder2
        for i in os.listdir(path):
              self.listWidget_2.insertItem(self.normal, i)
            
              self.normal = self.normal+1
        self.radioButton_7.setVisible(True)
        self.radioButton_7.setText("İsim değiştir") 
        
    def renamecovid(self):
        
        if(self.radioButton_6.isChecked() == True):
            self.lineEdit.setText("")
            self.groupBox_5.setVisible(True)
            
        if ( self.radioButton_6.isChecked() == False):
          
            self.groupBox_5.setVisible(False)
            
    def renamenormal(self):
        if(self.radioButton_7.isChecked() == True):
            self.lineEdit.setText("")
            self.groupBox_5.setVisible(True)
           
        if ( self.radioButton_7.isChecked() == False):
           
            self.groupBox_5.setVisible(False)        
            
            
    def rename(self):
        if self.radioButton_6.isChecked() == True & self.radioButton_7.isChecked()==True:
             QMessageBox.question(self, 'Bilgilendirme', "Aynı anda bir kaynak belirtin..",QMessageBox.Yes)  
             self.radioButton_6.setChecked(False)
             self.radioButton_7.setChecked(False)
        if (self.radioButton_7.isChecked() == False) and (self.radioButton_6.isChecked() == False):
             QMessageBox.question(self, 'Bilgilendirme', "Kaynak Belirtin..",QMessageBox.Yes) 
        if self.radioButton_6.isChecked()==True:
            if self.lineEdit.text() == "":
                QMessageBox.question(self, 'Bilgilendirme', "Değşitirilecek ismi yazınız..",QMessageBox.Yes) 
            else:
                sayac=0
               
                self.covidlist.clear()
                folder='./database/covid/'
                new = './databaserename/' 
                if not os.path.exists(new):
                    os.makedirs(new)
                newfile='./databaserename/covid/'
                if not os.path.exists(newfile):
                    os.makedirs(newfile)
                for i in os.listdir(folder):
                    name=self.lineEdit.text()+ str(sayac)
                    sayac=sayac+ 1
                    image = Image.open(folder+i)
                    new_image = image.resize((224, 224))
                    new_image.save(newfile+i)
                    os.rename(newfile+ i, newfile + name + ".jpg")
                self.listWidget.clear()
                for i in os.listdir(newfile):
                  self.listWidget.insertItem(self.recovid, i)
               
                  self.recovid = self.recovid+1
                self.folder1=newfile 
                print(self.folder1)
        if self.radioButton_7.isChecked()==True:
            if self.lineEdit.text() == "":
                QMessageBox.question(self, 'Bilgilendirme', "Değşitirilecek ismi yazınız..",QMessageBox.Yes) 
            else:
               
                sayac=0
                self.normallist.clear()
                folder='./database/normal/'
                new = './databaserename/' 
                if not os.path.exists(new):
                    os.makedirs(new)
                newfile='./databaserename/normal/'
                if not os.path.exists(newfile):
                    os.makedirs(newfile)
              
                for i in os.listdir(folder):
                    name=self.lineEdit.text()+ str(sayac)
                    sayac=sayac+ 1
                    image = Image.open(folder+i)
                    new_image = image.resize((224, 224))
                    new_image.save(newfile+i)
                    os.rename(newfile+ i, newfile + name + ".jpg")
                self.listWidget_2.clear()
                for i in os.listdir(newfile):
                  self.listWidget_2.insertItem(self.recovid, i)
                
                  self.renormal = self.renormal+1
                self.folder2=newfile 
                print(self.folder2)
                    
        
    def doubleclickcovid(self):
            self.label_6.setText("")
            item = self.listWidget.currentItem() 
            konum2='./yeniboyut/'
            if item is not None:
                 resimismi=item.text()
                 image = Image.open(self.folder1+resimismi)
                 new_image = image.resize((240, 240))
                 new_image.save(konum2+resimismi)
                 self.label_6.setStyleSheet("background-image : url('"+konum2+resimismi+"')") 
            
            
    def doubleclicknormal(self):
            self.label_6.setText("")
            item = self.listWidget_2.currentItem() 
            konum2='./yeniboyut/'
            if item is not None:
                 resimismi=item.text()
                 image = Image.open(self.folder2+resimismi)
                 new_image = image.resize((240, 240))
                 new_image.save(konum2+resimismi)
                 self.label_6.setStyleSheet("background-image : url('"+konum2+resimismi+"')")         
            
            
    def excelfile(self):
        databasename=self.lineEdit_2.text()
 
        isim_listesi=self.covidlist+self.normallist
        soyisim_listesi=self.covidnumber + self.normalnumber
        workbook = xlsxwriter.Workbook(databasename+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0,'id')
        worksheet.write(0,1,'patiens')
        worksheet.write(0,2,'target')
        s=1
        for i in range(0,len(isim_listesi)):
            
            worksheet.write(s,0,i)
            s=s+1
        for satir,veri in enumerate(isim_listesi):
            worksheet.write(satir+1,1,veri)
        for satir,veri in enumerate(soyisim_listesi):
            worksheet.write(satir+1,2,veri)
        workbook.close()
        
        veriseti=pd.read_excel('./'+databasename+'.xlsx')
        print(veriseti)
        self.data=pd.DataFrame(veriseti)
        islenmis=pandasModel(self.data)
        self.tableView.setModel(islenmis)
        
        data = pd.read_excel('./'+databasename+'.xlsx')
        print(data)   
        from sklearn.preprocessing import LabelEncoder
        from sklearn.compose import ColumnTransformer  
        le = LabelEncoder()
        data['target']= le.fit_transform(data['target']).astype(str)

        print(data)
        self.datax=pd.DataFrame(data)
        islenmis=pandasModel(self.datax)
        self.tableView_3.setModel(islenmis)
        
        y = pd.get_dummies(data.patiens)
        print(y)
        degisken=pd.DataFrame(y)
        islenmis=pandasModel(degisken)
        self.tableView_2.setModel(islenmis)
        

    def ninformation(self):
        n=int(self.lineEdit_3.text() ) 
        folder= './databaserename/covid/' 
        folder2='./databaserename/normal/' 
        for i in os.listdir(folder):
                    image = Image.open(folder+i)
                    new_image = image.resize((n, n))
                    new_image.save(folder+i)
            
        for i in os.listdir(folder2):
                    image = Image.open(folder2+i)
                    new_image = image.resize((n, n))
                    new_image.save(folder2+i)
                    
        self.label_11.setStyleSheet("color: darkblue; ")
        self.label_11.setText('Görüntülerin boyutları başarıyla değiştirildi.')



    def datacogalt(self):
        folder1array=[]
        folder2array=[]
        secim=self.comboBox.currentText()
        hedef=self.comboBox_2.currentText()
        count=int(self.lineEdit_6.text())
        if secim=='Flipping': # 180 derecelik çevirme işlemi yapar
            self.flippingfunction(hedef,count)
        elif secim== 'Rotation':#döndürme yapar
            self.rotationfunction(hedef, count)
        elif secim== 'Blur': #bulanıklaştırma
            self.blurfunction(hedef,count)
        else: #çerçeve ekleme
            self.makeborderfunction(hedef, count)
        pathdata='./databaserename/'
        folderc=pathdata+'covid/'
        foldern= pathdata+'normal/'
        for i in os.listdir(folderc):
            folder1array.append(i)
        for i in os.listdir(foldern):
            folder2array.append(i)
       
        
        self.label_2.setText("Datasetin içinde bulunan klasör: "+ 'covid'+"  Görüntü sayısı:"+str(len(folder1array)))
        self.label_3.setText("Datasetin içinde bulunan klasör: "+ 'covid' +"  Görüntü sayısı:"+str(len(folder2array)))
        
        self.listWidget.clear()
        
        for i in os.listdir(folderc):
              self.listWidget.insertItem(self.dacovid, i)
              self.covidlist.append(i)
              self.covidnumber.append("covid")
              self.dacovid = self.dacovid+1
        self.radioButton_6.setVisible(True)
        self.radioButton_6.setText("İsim değiştir")      	
        self.listWidget_2.clear()
        
        for i in os.listdir(foldern):
              self.listWidget_2.insertItem(self.danormal, i)
              self.normallist.append(i)
              self.normalnumber.append("normal")
              self.danormal = self.danormal+1
        self.radioButton_6.setVisible(True)
        self.radioButton_6.setText("İsim değiştir")  
    def flippingfunction(self, hedef, count):
        if hedef == 'Covid':
            path=self.folder1
        if hedef== 'Normal':
            path= self.folder2
        for i in os.listdir(path):
            name=" "
            count= count -1
            name=hedef+'DAF-'+str(count)
            newname=path+name+'.png'
            image = cv2.imread(path+i)
            x_axis = cv2.flip(image,0) #Vertical Flip
            cv2.imwrite(newname,x_axis)
          
            if count==0:
                break
           
    def rotationfunction(self, hedef, count):
        angle = 45
        if hedef == 'Covid':
            path=self.folder1
        if hedef== 'Normal':
            path= self.folder2
        for i in os.listdir(path):
            name=" "
            count= count -1
            name=hedef+'DAR-'+str(count)
            newname=path+name+'.png'
            image = cv2.imread(path+i)
            image_rotate = imutils.rotate_bound(image,angle)
            cv2.imwrite(newname,image_rotate)
           
            if count==0:
                break
        

    def blurfunction(self,hedef,count):
        if hedef == 'Covid':
            path=self.folder1
        if hedef== 'Normal':
            path= self.folder2
        for i in os.listdir(path):
            name=" "
            count= count -1
            name=hedef+'DAB-'+str(count)
            newname=path+name+'.png'
            image = cv2.imread(path+i)
            image_blur = cv2.GaussianBlur(image,(7,7),cv2.BORDER_DEFAULT)
            cv2.imwrite(newname,image_blur)
           
            if count==0:
                break
        
        
    def makeborderfunction(self,hedef,count):
        if hedef == 'Covid':
            path=self.folder1
        if hedef== 'Normal':
            path= self.folder2
        for i in os.listdir(path):
            name=" "
            count= count -1
            name=hedef+'DAB-'+str(count)
            newname=path+name+'.png'
            image = cv2.imread(path+i)
            image_border = cv2.copyMakeBorder(image, 20, 20, 20, 20, cv2.BORDER_CONSTANT)
            cv2.imwrite(newname,image_border)
           
            if count==0:
                break
          
    def datasplitfunction(self):
        databasename=self.lineEdit_2.text()
        self.comboBox_4.clear()
        self.kfoldlistesi=[]
      
        secim=self.comboBox_3.currentText()
        if secim=='Holdout':
            self.groupBox_16.setVisible(False)
            self.tableView_5.clearSpans()
            self.tableView_6.clearSpans()
            self.tableView_7.clearSpans()
            self.tableView_4.clearSpans()
            
            testsize=self.lineEdit_5.text()
            testsize=int(testsize)
            testsize=float(testsize/100)
            dataframe = pd.read_excel('./'+databasename+'.xlsx')
            array = dataframe.values 
            satir,sutun=dataframe.shape
            X = array[:,0:sutun-1] #bagımsız degiskenler
            y= array[:,sutun-1] #bagımlı degiskenler
            #verilerin egitim ve test icin bolunmesi
            self.x_train, self.x_test,self.y_train,self.y_test = train_test_split(X,y,test_size=testsize, random_state=0)

            degisken=pd.DataFrame(self.x_train)
            islenmis=pandasModel(degisken)
            self.tableView_4.setModel(islenmis)
            degisken=pd.DataFrame(self.y_train)
            islenmis=pandasModel(degisken)
            self.tableView_5.setModel(islenmis)
            degisken=pd.DataFrame(self.x_test)
            islenmis=pandasModel(degisken)
            self.tableView_6.setModel(islenmis)
            degisken=pd.DataFrame(self.y_test)
            islenmis=pandasModel(degisken)
            self.tableView_7.setModel(islenmis)
            print(self.x_train)
            print(self.x_test)
            print(self.y_test)
            print(self.y_train)
            self.splitdata()
        if secim=='KFold':
            self.groupBox_16.setVisible(True)
            self.tableView_7.clearSpans()
            self.tableView_4.clearSpans()
            self.tableView_5.clearSpans()
            self.tableView_6.clearSpans()
            
            testsize=self.lineEdit_5.text()
            testsize=int(testsize)
            self.comboBox_4.clear()
            for i in range(0,testsize):
                
                self.comboBox_4.addItem(str(i+1))

            dataframe = pd.read_excel('./'+databasename+'.xlsx')
            array = dataframe.values 
            satir,sutun=dataframe.shape
            X = array[:,0:sutun-1] #bagımsız degiskenler
            y= array[:,sutun-1] #bagımlı degiskenler
            i=0
            cv = KFold(n_splits=testsize, random_state = 0, shuffle=True)
            for train_index, test_index in cv.split(X):
       
                self.x_train, self.x_test, self.y_train, self.y_test = X[train_index], X[test_index], y[train_index], y[test_index]
                self.kfoldlistesi.append(self.x_train)
                self.kfoldlistesi.append(self.x_test)
                self.kfoldlistesi.append(self.y_train)
                self.kfoldlistesi.append(self.y_test)
            print(self.x_train)
            print(self.x_test)
            print(self.y_test)
            print(self.y_train)
                  
    def splitdata(self):
          
                folder='./database/normal/'
                folder2='./database/covid/'
                new = './mydata/' 
                if not os.path.exists(new):
                    os.makedirs(new)
                newfile='./mydata/normal/'
                newfile2='./mydata/covid'
                if not os.path.exists(newfile):
                    os.makedirs(newfile)
                if not os.path.exists(newfile2):
                    os.makedirs(newfile2)
                    
                sayac=0   
                for i in os.listdir(folder):
                    name=self.lineEdit.text()+ str(sayac)
                    sayac=sayac+ 1
                    image = Image.open(folder+i)
                    new_image = image.resize((224, 224))
                    new_image.save(newfile+i)
                    os.rename(newfile+ i, newfile + name + ".jpg")
        
    def kfoldyapalim(self):
        secim=self.comboBox_4.currentText()
        if secim =="1":
           degisken=self.kfoldlistesi[0]
           degisken1=self.kfoldlistesi[1]
           degisken2=self.kfoldlistesi[2]
           degisken3=self.kfoldlistesi[3] 
        if secim=="2":
           degisken=self.kfoldlistesi[4]
           degisken1=self.kfoldlistesi[5]
           degisken2=self.kfoldlistesi[6]
           degisken3=self.kfoldlistesi[7]
        if secim=="3":
           degisken=self.kfoldlistesi[8]
           degisken1=self.kfoldlistesi[9]
           degisken2=self.kfoldlistesi[10]
           degisken3=self.kfoldlistesi[11]      
        if secim=="4":
           degisken=self.kfoldlistesi[12]
           degisken1=self.kfoldlistesi[13]
           degisken2=self.kfoldlistesi[14]
           degisken3=self.kfoldlistesi[15]
        if secim=="5":
           degisken=self.kfoldlistesi[16]
           degisken1=self.kfoldlistesi[17]
           degisken2=self.kfoldlistesi[18]
           degisken3=self.kfoldlistesi[19]
        degisken=pd.DataFrame(degisken)
        islenmis=pandasModel(degisken)
        self.tableView_4.setModel(islenmis)
        degisken2=pd.DataFrame(degisken2)
        islenmis=pandasModel(degisken2)
        self.tableView_5.setModel(islenmis)
        degisken1=pd.DataFrame(degisken1)
        islenmis=pandasModel(degisken1)
        self.tableView_6.setModel(islenmis)
        degisken3=pd.DataFrame(degisken3)
        islenmis=pandasModel(degisken3)
        self.tableView_7.setModel(islenmis)     


    def test(self, item):
    
        d=item.data()
        print(d)
        newsize='./yeniboyut/'
        fldr='./databaserename/covid/'
        fldr2='./databaserename/normal/'
        for i in os.listdir(fldr):
            if i==d:
                
                image = Image.open(fldr+i)
                new_image = image.resize((260, 260))
                new_image.save(newsize+i)
                self.label_22.setStyleSheet("background-image : url('"+newsize+i+"')")   
                self.label_23.setText(d)
        for i in os.listdir(fldr2):
            if i==d:
                
                image = Image.open(fldr2+i)
                new_image = image.resize((260, 260))
                new_image.save(newsize+i)
                self.label_22.setStyleSheet("background-image : url('"+newsize+i+"')")   
                self.label_23.setText(d)




  
    def transferlearn(self):
        
        if self.radioButton.isChecked():
                self.vgg16funciton()
        if self.radioButton_3.isChecked():
            self.inceptionfunction()
        if self.radioButton_2.isChecked():
            self.resnetfunction()
        if self.radioButton_4.isChecked():
            self.mobilenetfunction()
        if self.radioButton_5.isChecked():
            self.xceptionfunction()

       


    def vgg16funciton(self):
                imgsize=int(self.lineEdit_3.text())
                from keras.preprocessing.image import ImageDataGenerator
                train_datagen = ImageDataGenerator(rescale = 1./255,
                                                   shear_range = 0.2,
                                                   zoom_range = 0.2,
                                                   validation_split=0.2,
                                                   horizontal_flip = True)
                test_datagen = ImageDataGenerator(rescale = 1./255)
                training_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')
                validation_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')
                test_set = test_datagen.flow_from_directory(r'./data/test/', target_size = (imgsize, imgsize), class_mode = 'categorical')
                IMAGE_SIZE = [imgsize, imgsize]
                vgg = VGG16(input_shape=IMAGE_SIZE + [3], weights='imagenet', include_top=False)
                for layer in vgg.layers:
                 layer.trainable = False
                x = Flatten()(vgg.output)
                prediction = Dense(1, activation='sigmoid')(x)
                model = Model(inputs=vgg.input, outputs=prediction)
                model.compile(loss='categorical_crossentropy',
                                    optimizer=optimizers.Adam(),
                                    metrics=['accuracy'])
                model.summary()    
                from datetime import datetime
                from keras.callbacks import ModelCheckpoint, LearningRateScheduler
                from keras.callbacks import ReduceLROnPlateau
                lr_reducer = ReduceLROnPlateau(factor=np.sqrt(0.1),
                                               cooldown=0,
                                               patience=5,
                                               min_lr=0.5e-6)
                checkpoint = ModelCheckpoint(filepath='mymodel.h5', 
                                               verbose=1, save_best_only=True)
                callbacks = [checkpoint, lr_reducer]
                start = datetime.now()
                history = model.fit_generator(training_set, 
                                    steps_per_epoch=1, 
                                    epochs = 3, verbose=1, 
                                    validation_data = validation_set, 
                                    validation_steps = 1)
                duration = datetime.now() - start
                score = model.evaluate(test_set)
                print('Test Loss:', score[0])
                print('Test accuracy:', score[1])
                testloss=round(score[0],2)
                testacc=round(score[1],2)
                self.label_16.setText('Model eğitim sonuçları:'+'\n'+'Test Loss:'+ str(testloss)+'\n'+'Test accuracy:'+ str(testacc))
                
                import matplotlib.pyplot as plt
                plt.figure()
                plt.plot(history.history["accuracy"])
                plt.plot(history.history["loss"])
                plt.title("model accuracy")
                plt.ylabel("Accuracy")
                plt.xlabel("Epoch")
                plt.legend(["Accuracy","loss"])       
                plt.savefig('./vggsonuc',dpi=300)
                plt.show();
                image = Image.open('./vggsonuc.png')
                new_image = image.resize((250, 250))
                new_image.save('./vggsonuc.png')
                self.label_15.setStyleSheet("background-image : url('./vggsonuc.png')")
                
                from sklearn import metrics
                preds = model.predict(validation_set, verbose=1)
                y_pred = np.argmax(preds, axis=1)
                cnf_matrix = confusion_matrix(validation_set.classes, y_pred)
                np.set_printoptions(precision=2)
                degis=metrics.classification_report(validation_set.classes, y_pred, labels=[0,1])
                print(degis)
                self.label_24.setText(str(degis))
                print(cnf_matrix)
            
                fpr, tpr, _ = roc_curve(validation_set.classes, preds)
                roc_auc = auc(fpr, tpr)
                plt.figure()
                lw = 2
                plt.plot(fpr, tpr, color='darkred',
                           lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
                plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
                plt.xlim([0.0, 1.0])
                plt.ylim([0.0, 1.05])
                plt.xlabel('False Positive Rate')
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic')
                plt.legend(loc="lower right")
                plt.savefig('./vggroc',dpi=300)
                image = Image.open('./vggroc.png')
                new_image = image.resize((250, 250))
                new_image.save('./vggroc.png')
                self.label_20.setStyleSheet("background-image : url('./vggroc.png')")
                model_json = model.to_json()
                with open("./vgg16.json", "w") as json_file:
                    json_file.write(model_json)
                
                model.save_weights("./vgg16.h5")
                print("Saved model to disk")
                self.label_9.setStyleSheet('color:darkred')
                self.label_9.setText("Accuracy-Loss") 
                self.label_21.setStyleSheet('color:darkred')
                self.label_21.setText("ROC Curve") 
                
                from openpyxl import load_workbook
                workbook = load_workbook(filename="modellog.xlsx")
                sheet = workbook.active
                sheet["B2"] = 'VGG16'
                sheet["C2"] = str(testacc)
                sheet["D2"] = str(testloss)
                workbook.save(filename="modellog.xlsx");
                self.modelresult()
    def inceptionfunction(self):
           imgsize=int(self.lineEdit_3.text())
           from keras.preprocessing.image import ImageDataGenerator
           train_datagen = ImageDataGenerator(rescale = 1./255,
                                                   shear_range = 0.2,
                                                   zoom_range = 0.2,
                                                   validation_split=0.2,
                                                   horizontal_flip = True)
            
           test_datagen = ImageDataGenerator(rescale = 1./255)
           training_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')   
           validation_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')  
           test_set = test_datagen.flow_from_directory(r'./data/test/', target_size = (imgsize, imgsize), class_mode = 'categorical')
           from keras.models import Sequential
           from keras.models import Model
           from keras.callbacks import ModelCheckpoint, LearningRateScheduler, EarlyStopping, ReduceLROnPlateau, TensorBoard
           from keras import optimizers, losses, activations, models
           from keras.layers import Convolution2D, Dense, Input, Flatten, Dropout, MaxPooling2D, BatchNormalization, GlobalAveragePooling2D, Concatenate
           from keras import applications
          #önceden eğitilmiş bi ağ varsa onu donduruyor
           base_model = applications.InceptionV3(weights='imagenet', 
                                            include_top=False, 
                                            input_shape=(imgsize, imgsize,3))
           base_model.trainable = False
           add_model = Sequential()
           add_model.add(base_model)
           add_model.add(GlobalAveragePooling2D())
           add_model.add(Dropout(0.5))
           add_model.add(Dense(1, activation='sigmoid'))
           model = add_model
           model.compile(loss='categorical_crossentropy',  optimizer=optimizers.SGD(lr=1e-4, momentum=0.9), metrics=['accuracy'])
           model.summary()
           file_path="./inceptionweights.best.hdf5"
           checkpoint = ModelCheckpoint(file_path, monitor='acc', verbose=1, save_best_only=True, mode='max')
           early = EarlyStopping(monitor="acc", mode="max", patience=3)
           callbacks_list = [checkpoint, early] #early
           history = model.fit_generator(training_set, 
                                          epochs=2, 
                                          shuffle=True, 
                                          verbose=True,
                                          callbacks=callbacks_list)     
          
           scorea = model.evaluate(test_set)
           print('Test Loss:', scorea[0])
           print('Test accuracy:', scorea[1])
           testloss=round(scorea[0],2)
           testacc=round(scorea[1],2)
           self.label_16.setText('Model eğitim sonuçları:'+'\n'+'Test Loss:'+ str(testloss)+'\n'+'Test accuracy:'+ str(testacc))
           plt.figure()
           plt.plot(history.history["accuracy"])
           plt.plot(history.history["loss"])
           plt.title("model accuracy")
           plt.ylabel("Accuracy")
           plt.xlabel("Loss")
           plt.legend(["Accuracy","loss"])
           plt.savefig('./inceptionsonuc',dpi=300) 
           image = Image.open('./inceptionsonuc.png')
           new_image = image.resize((250, 250))
           new_image.save('./inceptionsonuc.png')
           self.label_15.setStyleSheet("background-image : url('./inceptionsonuc.png')")
           
           preds = model.predict(validation_set,verbose=1)
           y_pred = np.argmax(preds, axis=1)
           cnf_matrix = confusion_matrix(validation_set.classes, y_pred)
           np.set_printoptions(precision=2)
           degis=metrics.classification_report(validation_set.classes, y_pred, labels=[0,1])
           print(degis)
           self.label_24.setText(str(degis))
           print(cnf_matrix)
            
           fpr, tpr, _ = roc_curve(validation_set.classes, preds)
           roc_auc = auc(fpr, tpr)
           plt.figure()
           lw = 2
           plt.plot(fpr, tpr, color='darkred',
                     lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
           plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
           plt.xlim([0.0, 1.0])
           plt.ylim([0.0, 1.05])
           plt.xlabel('False Positive Rate')
           plt.ylabel('True Positive Rate')
           plt.title('Receiver Operating Characteristic')
           plt.legend(loc="lower right")
           plt.savefig('./inceptionroc',dpi=300)
           image = Image.open('./inceptionroc.png')
           new_image = image.resize((250, 250))
           new_image.save('./inceptionroc.png')
           self.label_20.setStyleSheet("background-image : url('./inceptionroc.png')")
           model_json = model.to_json()
           with open("./inceptionv3.json", "w") as json_file:
                    json_file.write(model_json)
                # serialize weights to HDF5
           model.save_weights("./inceptionv3.h5")
           print("Saved model to disk")
           self.label_9.setStyleSheet('color:darkred')
           self.label_9.setText("Accuracy-Loss") 
           self.label_21.setStyleSheet('color:darkred')
           self.label_21.setText("ROC Curve") 
           
           from openpyxl import load_workbook
           workbook = load_workbook(filename="modellog.xlsx")
           sheet = workbook.active
           sheet["B3"] = 'Inception V3'
           sheet["C3"] = str(testacc)
           sheet["D3"] = str(testloss)  
           workbook.save(filename="modellog.xlsx");
           self.modelresult()
    def resnetfunction(self):
            from keras.layers import Convolution2D,GlobalAveragePooling2D
            from keras.layers import Convolution2D, Dense, Input, Flatten, Dropout
            imgsize=int(self.lineEdit_3.text())
            from keras.applications.resnet50 import ResNet50
            from keras.applications.resnet50 import preprocess_input
            from keras.preprocessing.image import ImageDataGenerator
 
            train_datagen = ImageDataGenerator(rescale = 1./255,
                                                   shear_range = 0.2,
                                                   zoom_range = 0.2,
                                                   validation_split=0.2,
                                                   horizontal_flip = True)
            
            test_datagen = ImageDataGenerator(rescale = 1./255)
            training_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')
            validation_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical') 
            test_set = test_datagen.flow_from_directory(r'./data/test/', target_size = (imgsize, imgsize), class_mode = 'categorical') 
            
            model = Sequential()
            model.add(ResNet50(include_top = False, weights="imagenet", pooling = 'avg', input_shape=(imgsize,imgsize,3)))
            model.add(Dense(512, activation='relu', input_dim=(imgsize,imgsize,3)))
            model.add(Dropout(0.3))
            model.add(Dense(512, activation='relu'))
            model.add(Dropout(0.3))
            model.add(Dense(1, activation = 'sigmoid'))
            model.layers[0].trainable = False
            model.summary()
            sgd = optimizers.SGD(lr = 0.01, decay = 1e-6, momentum = 0.9, nesterov = True)
            model.compile(optimizer = sgd, loss = 'categorical_crossentropy', metrics = ['accuracy'])
            from tensorflow.python.keras.callbacks import EarlyStopping, ModelCheckpoint
            cb_early_stopper = EarlyStopping(monitor = 'val_loss', patience = 10)
            cb_checkpointer = ModelCheckpoint( filepath='mymodel.h5', monitor = 'val_loss', save_best_only = True, mode = 'auto'   )
            fit_history = model.fit_generator(
            training_set,
            steps_per_epoch=1,
            epochs = 2,
            validation_data=validation_set,
            validation_steps=1,
            callbacks=[cb_checkpointer, cb_early_stopper] ) 
            scorea = model.evaluate(test_set)
            print('Test Loss:', scorea[0])
            print('Test accuracy:', scorea[1])
            testloss=round(scorea[0],2)
            testacc=round(scorea[1],2)
            self.label_16.setText('Model eğitim sonuçları:'+'\n'+'Test Loss:'+ str(testloss)+'\n'+'Test accuracy:'+ str(testacc))
            plt.figure()
            plt.plot(fit_history.history["loss"])
            plt.plot(fit_history.history["accuracy"])
            plt.title("model accuracy")
            plt.ylabel("Accuracy")
            plt.xlabel("Loss")
            plt.legend(["Accuracy","loss"])
            plt.savefig('./resnetsonuc',dpi=300) 
            plt.show();
            image = Image.open('./resnetsonuc.png')
            new_image = image.resize((250, 250))
            new_image.save('./resnetsonuc.png')
            self.label_15.setStyleSheet("background-image : url('./resnetsonuc.png')")
            
            preds = model.predict(validation_set,
                                  verbose=1)
            y_pred = np.argmax(preds, axis=1)
            cnf_matrix = confusion_matrix(validation_set.classes, y_pred)
            np.set_printoptions(precision=2)
            degis=metrics.classification_report(validation_set.classes, y_pred, labels=[0,1])
            print(degis)
            self.label_24.setText(str(degis))
            print(cnf_matrix)
            fpr, tpr, _ = roc_curve(validation_set.classes, preds)
            roc_auc = auc(fpr, tpr)
            plt.figure()
            lw = 2
            plt.plot(fpr, tpr, color='darkred',
                     lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
            plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
            plt.xlim([0.0, 1.0])
            plt.ylim([0.0, 1.05])
            plt.xlabel('False Positive Rate')
            plt.ylabel('True Positive Rate')
            plt.title('Receiver Operating Characteristic')
            plt.legend(loc="lower right")
            plt.savefig('./resnetroc',dpi=300)
            image = Image.open('./resnetroc.png')
            new_image = image.resize((250, 250))
            new_image.save('./resnetroc.png')
            self.label_20.setStyleSheet("background-image : url('./resnetroc.png')")
            model_json = model.to_json()
            with open("./resnet50.json", "w") as json_file:
                json_file.write(model_json)
            model.save_weights("./resnet50.h5")
            print("Saved model to disk")
            self.label_9.setStyleSheet('color:darkred')
            self.label_9.setText("Accuracy-Loss") 
            self.label_21.setStyleSheet('color:darkred')
            self.label_21.setText("ROC Curve") 
            from openpyxl import load_workbook
            workbook = load_workbook(filename="modellog.xlsx")
            sheet = workbook.active
                
                # Write what you want into a specific cell
            sheet["B4"] = 'ResNet 50'
            sheet["C4"] = str(testacc)
            sheet["D4"] = str(testloss)
                
                
                # Save the spreadsheet
            workbook.save(filename="modellog.xlsx");
            self.modelresult()
        
    def mobilenetfunction(self):
            imgsize=int(self.lineEdit_3.text())
            from keras.preprocessing.image import ImageDataGenerator
    
            train_datagen = ImageDataGenerator(rescale = 1./255,
                                                   shear_range = 0.2,
                                                   zoom_range = 0.2,
                                                   validation_split=0.2,
                                                   horizontal_flip = True)
            
            test_datagen = ImageDataGenerator(rescale = 1./255)

            training_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')
            validation_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (imgsize, imgsize),class_mode = 'categorical')   
            test_set = test_datagen.flow_from_directory(r'./data/test/', target_size = (imgsize, imgsize), class_mode = 'categorical')
            model = Sequential()
            model.add(MobileNetV2(include_top = False, weights="imagenet", input_shape=(imgsize, imgsize, 3)))
            model.add(tf.keras.layers.GlobalAveragePooling2D())
            model.add(Dense(1, activation = 'sigmoid'))
            model.layers[0].trainable = False
            model.compile(optimizer=RMSprop(lr=0.01), loss = 'categorical_crossentropy', metrics = 'accuracy')
            history = model.fit(training_set,
                steps_per_epoch=1,
                epochs=3,
                verbose=1,
                validation_data = validation_set,
                validation_steps=1)       
            model.evaluate(validation_set)
            score = model.evaluate(test_set)
            print('Test Loss:', score[0])
            print('Test accuracy:', score[1])
            testloss=round(score[0],2)
            testacc=round(score[1],2)
            self.label_16.setText('Model eğitim sonuçları:'+'\n'+'Test Loss:'+ str(testloss)+'\n'+'Test accuracy:'+ str(testacc))
            import matplotlib.pyplot as plt
            plt.plot(history.history["accuracy"])  
            plt.plot(history.history["loss"])
            plt.title("model accuracy")
            plt.ylabel("Accuracy")
            plt.xlabel("Loss")
            plt.legend(["Accuracy","loss"])
            plt.savefig('./mobilenetv2',dpi=300)
            image = Image.open('./mobilenetv2.png')
            new_image = image.resize((250, 250))
            new_image.save('./mobilenetv2.png')
            self.label_15.setStyleSheet("background-image : url('./mobilenetv2.png')")
            
            preds = model.predict(validation_set,
                                  verbose=1)
           
            y_pred = np.argmax(preds, axis=1)
            cnf_matrix = confusion_matrix(validation_set.classes, y_pred)
            np.set_printoptions(precision=2)
            degis=metrics.classification_report(validation_set.classes, y_pred, labels=[0,1])
            print(degis)
            self.label_24.setText(str(degis))
            print(cnf_matrix)
            fpr, tpr, _ = roc_curve(validation_set.classes, preds)
            roc_auc = auc(fpr, tpr)
            plt.figure()
            lw = 2
            plt.plot(fpr, tpr, color='darkred',
                     lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
            plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
            plt.xlim([0.0, 1.0])
            plt.ylim([0.0, 1.05])
            plt.xlabel('False Positive Rate')
            plt.ylabel('True Positive Rate')
            plt.title('Receiver Operating Characteristic')
            plt.legend(loc="lower right")
            plt.savefig('./mobilenetv2roc',dpi=300)
            image = Image.open('./mobilenetv2roc.png')
            new_image = image.resize((250, 250))
            new_image.save('./mobilenetv2roc.png')
            self.label_20.setStyleSheet("background-image : url('./mobilenetv2roc.png')")
            model_json = model.to_json()
            with open("./mobilenet.json", "w") as json_file:
                    json_file.write(model_json)
                # serialize weights to HDF5
            model.save_weights("./mobilenet.h5")
            print("Saved model to disk")
            self.label_9.setStyleSheet('color:darkred')
            self.label_9.setText("Accuracy-Loss") 
            self.label_21.setStyleSheet('color:darkred')
            self.label_21.setText("ROC Curve") 
            from openpyxl import load_workbook
            workbook = load_workbook(filename="modellog.xlsx")
            sheet = workbook.active
                
                # Write what you want into a specific cell
            sheet["B5"] = 'MobileNet V2'
            sheet["C5"] = str(testacc)
            sheet["D5"] = str(testloss)
                
                
                # Save the spreadsheet
            workbook.save(filename="modellog.xlsx");
            self.modelresult()
    def xceptionfunction(self):
            from tf_explain.core.activations import ExtractActivations
            from tensorflow.keras.applications.xception import decode_predictions
            from keras.applications.resnet50 import preprocess_input
            from keras.preprocessing.image import ImageDataGenerator
            train_datagen = ImageDataGenerator(rescale = 1./255,
                                                   shear_range = 0.2,
                                                   zoom_range = 0.2,
                                                   validation_split=0.2,
                                                   horizontal_flip = True)
            
            test_datagen = ImageDataGenerator(rescale = 1./255)
            training_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (220, 220),class_mode = 'categorical')
                
            validation_set = train_datagen.flow_from_directory(r'./data/train/', target_size = (220, 220),class_mode = 'categorical')
                
            test_set = test_datagen.flow_from_directory(r'./data/test/', target_size = (220, 220), class_mode = 'categorical')
        
            base_model = tf.keras.applications.xception.Xception(input_shape=(220, 220, 3), weights='imagenet', include_top=False)
        
            # Top Model Block
            x = base_model.output
          
            predictions = Dense(1, activation='sigmoid')(x)
        
            # add your top layer block to your base model
            model = Model(base_model.input, predictions)
             
            model.layers[0].trainable = False
            model.summary()
            
            sgd = optimizers.SGD(lr = 0.01, decay = 1e-6, momentum = 0.9, nesterov = True)
            model.compile(optimizer = sgd, loss = 'categorical_crossentropy', metrics = ['accuracy'])
            from tensorflow.python.keras.callbacks import EarlyStopping, ModelCheckpoint
    
            cb_early_stopper = EarlyStopping(monitor = 'val_loss', patience = 10)
            cb_checkpointer = ModelCheckpoint(filepath = 'best.hdf5', monitor = 'val_loss', save_best_only = True, mode = 'auto'   )
            fit_history = model.fit_generator(
            training_set,
            steps_per_epoch=1,
            epochs = 2,
            validation_data=validation_set,
            validation_steps=1,
            callbacks=[cb_checkpointer, cb_early_stopper]
            )
            scorea = model.evaluate(test_set)
            print('Test Loss:', scorea[0])
            print('Test accuracy:', scorea[1])
            testloss=round(scorea[0],2)
            testacc=round(scorea[1],2)
            self.label_16.setText('Model eğitim sonuçları:'+'\n'+'Test Loss:'+ str(testloss)+'\n'+'Test accuracy:'+ str(testacc))
            plt.figure()
            plt.plot(fit_history.history["accuracy"])
            
            plt.plot(fit_history.history["loss"])
          
            plt.title("model accuracy")
            plt.ylabel("Accuracy")
            plt.xlabel("Epoch")
            plt.legend(["Accuracy","loss"])
           
                
            plt.savefig('./xceptionsonuc.png',dpi=300) 
            plt.show();
            image = Image.open('./xceptionsonuc.png')
            new_image = image.resize((250, 250))
            new_image.save('./xceptionsonuc.png')
            self.label_15.setStyleSheet("background-image : url('./xceptionsonuc.png')")
            preds = model.predict(validation_set,
                                  verbose=1)
            fpr, tpr, _ = roc_curve(validation_set.classes, preds)
            roc_auc = auc(fpr, tpr)
            plt.figure()
            lw = 2
            plt.plot(fpr, tpr, color='darkred',
                     lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
            plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
            plt.xlim([0.0, 1.0])
            plt.ylim([0.0, 1.05])
            plt.xlabel('False Positive Rate')
            plt.ylabel('True Positive Rate')
            plt.title('Receiver Operating Characteristic')
            plt.legend(loc="lower right")
            plt.savefig('./xceptionroc',dpi=300)
            image = Image.open('./xceptionroc.png')
            new_image = image.resize((250, 250))
            new_image.save('./xceptionroc.png')
            self.label_20.setStyleSheet("background-image : url('./xceptionroc.png')")
            model_json = model.to_json()
            with open("./xception.json", "w") as json_file:
                    json_file.write(model_json)
                # serialize weights to HDF5
            model.save_weights("./xception.h5")
            print("Saved model to disk")
            self.label_9.setStyleSheet('color:darkred')
            self.label_9.setText("Accuracy-Loss") 
            self.label_21.setStyleSheet('color:darkred')
            self.label_21.setText("ROC Curve") 
            from openpyxl import load_workbook
            workbook = load_workbook(filename="modellog.xlsx")
            sheet = workbook.active
                
                
            sheet["B6"] = 'XCeption'
            sheet["C6"] = str(testacc)
            sheet["D6"] = str(testloss)
                
                
               
            workbook.save(filename="modellog.xlsx");
            self.modelresult()
    def modeltestyap(self):
        self.dosyaadi, isim=QtWidgets.QFileDialog.getOpenFileName(None,"Lütfen tahmin edilecek görüntüyü seçin.","","Veri Seti Türü(*.jpg)")
        self.label_17.setText(self.dosyaadi[80:])
        image = Image.open(self.dosyaadi)
        new_image = image.resize((240, 240))
        new_image.save(self.dosyaadi)
        self.label_18.setStyleSheet("background-image : url('"+self.dosyaadi+"')")
        if(self.dosyaadi == None):
            QMessageBox.question(self, 'Bilgilendirme', "İşlenecek resim seçilmedi.",QMessageBox.Yes)
        else:
            if self.radioButton.isChecked():
                self.vgg16model()
            if self.radioButton_3.isChecked():
                self.inceptionmodel()
            if self.radioButton_2.isChecked():
                self.resnetmodel()
            if self.radioButton_4.isChecked():
                self.mobilenetmodel()
            if self.radioButton_5.isChecked():
                self.xceptionmodel()

         
        
          
        
            
    def vgg16model(self):
            json_file = open('./vgg16.json', 'r')
            loaded_model_json = json_file.read()
            json_file.close()
            self.loaded_model = model_from_json(loaded_model_json)
            # load weights into new model
            self.loaded_model.load_weights("./vgg16.h5")
            print("Loaded model from disk") 
            image = Image.open(self.dosyaadi)
            new_image = image.resize((224, 224))
            new_image.save(self.dosyaadi)
            
            test_image = Image.open(self.dosyaadi)
            test_image = img_to_array(test_image)
          
            test_image = np.expand_dims(test_image, axis = 0)
            result = self.loaded_model.predict(test_image)
           
            res = np.argmax(result)
            
            self.label_19.setStyleSheet('color:darkred')
            print(result[0])
            if result[0]<0.5:
                self.label_19.setText("Modelin Tahmini: Covid")       
            else:
                self.label_19.setText("Modelin Tahmini: Not Covid")  
    def resnetmodel(self):
            json_file = open('./resnet50.json', 'r')
            loaded_model_json = json_file.read()
            json_file.close()
            self.loaded_model = model_from_json(loaded_model_json)
            # load weights into new model
            self.loaded_model.load_weights("./resnet50.h5")
            print("Loaded model from disk") 
            image = Image.open(self.dosyaadi)
            new_image = image.resize((224, 224))
            new_image.save(self.dosyaadi)
            
            test_image = Image.open(self.dosyaadi)
            test_image = img_to_array(test_image)
          
            test_image = np.expand_dims(test_image, axis = 0)
            result = self.loaded_model.predict(test_image)
           
            res = np.argmax(result)
            
            self.label_19.setStyleSheet('color:darkred')
            print(result[0])
            if result[0]==0:
                self.label_19.setText("Modelin Tahmini: Covid")       
            else:
                self.label_19.setText("Modelin Tahmini: Not Covid")  
    def inceptionmodel(self):   
            json_file = open('./inceptionv3.json', 'r')
            loaded_model_json = json_file.read()
            json_file.close()
            self.loaded_model = model_from_json(loaded_model_json)
            # load weights into new model
            self.loaded_model.load_weights("./inceptionv3.h5")
            print("Loaded model from disk")
            image = Image.open(self.dosyaadi)
            new_image = image.resize((224, 224))
            new_image.save(self.dosyaadi)
            
            test_image = Image.open(self.dosyaadi)
            test_image = img_to_array(test_image)
          
            test_image = np.expand_dims(test_image, axis = 0)
            result = self.loaded_model.predict(test_image)
           
            res = np.argmax(result)
            
            self.label_19.setStyleSheet('color:darkred')
            print(result[0])
            if result[0]<0.5:
                self.label_19.setText("Modelin Tahmini: Covid")       
            else:
                self.label_19.setText("Modelin Tahmini: Not Covid")  
    def xceptionmodel(self):
            json_file = open('./xception.json', 'r')
            loaded_model_json = json_file.read()
            json_file.close()
            self.loaded_model = model_from_json(loaded_model_json)
            # load weights into new model
            self.loaded_model.load_weights("./xception.h5")
            print("Loaded model from disk")
            image = Image.open(self.dosyaadi)
            new_image = image.resize((224, 224))
            new_image.save(self.dosyaadi)
            
            test_image = Image.open(self.dosyaadi)
            test_image = img_to_array(test_image)
          
            test_image = np.expand_dims(test_image, axis = 0)
            result = self.loaded_model.predict(test_image)
           
            res = np.argmax(result)
            
            self.label_19.setStyleSheet('color:darkred')
            print(result[0])
            if result[0]==0:
                self.label_19.setText("Modelin Tahmini: Covid")       
            else:
                self.label_19.setText("Modelin Tahmini: Not Covid")  
    def mobilenetmodel(self):
            json_file = open('./mobilenet.json', 'r')
            loaded_model_json = json_file.read()
            json_file.close()
            self.loaded_model = model_from_json(loaded_model_json)
            # load weights into new model
            self.loaded_model.load_weights("./mobilenet.h5")
            print("Loaded model from disk")
            image = Image.open(self.dosyaadi)
            new_image = image.resize((224, 224))
            new_image.save(self.dosyaadi)
            
            test_image = Image.open(self.dosyaadi)
            test_image = img_to_array(test_image)
          
            test_image = np.expand_dims(test_image, axis = 0)
            result = self.loaded_model.predict(test_image)
           
            res = np.argmax(result)
            
            self.label_19.setStyleSheet('color:darkred')
            print(result[0])
            if result[0]<0.5:
                self.label_19.setText("Modelin Tahmini: Covid")       
            else:
                self.label_19.setText("Modelin Tahmini: Not Covid")  
   
    def modelresult(self):
        veriseti=pd.read_excel('./modellog.xlsx')
       
        self.data=pd.DataFrame(veriseti)
        islenmis=pandasModel(self.data)
        self.tableView_8.setModel(islenmis)
        

class pandasModel(QAbstractTableModel):
    def __init__(self, data):
        QAbstractTableModel.__init__(self)
        self._data = data
    def rowCount(self, parent=None):
        return self._data.shape[0]
    def columnCount(self, parnet=None):
        return self._data.shape[1]
        
    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])
        return None
    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._data.columns[col]
        return None
                
                            